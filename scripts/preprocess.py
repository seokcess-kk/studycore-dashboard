# -*- coding: utf-8 -*-
"""
inout_log.xlsx (이벤트 로그) + student-info.xlsx (명부) -> web/data.js

독서실 입·퇴장 '이벤트 로그'를 읽어 학생별/일별/월별 순공부시간·출결을 집계한다.
구버전은 시스템이 미리 집계한 세션 파일(inout_raw.xlsx)을 받았으나, 이제는
입장/외출/재입장/이동/퇴장/강제퇴장 단위의 원시 이벤트를 받아 영업일(새벽 5시
경계)로 묶고 구간을 직접 계산한다. 입·퇴 시각이 모두 raw로 남아 있어
자동퇴장 수동보정이 필요 없다(자정 강제퇴장 후 새벽 공부도 그대로 이어 계산).

이벤트 상태:
  입장·재입장          → 공부 구간 시작
  외출·이동            → 자리 비움(제외) 시작 + 직전 공부 구간 종료   ※ '이동'도 제외 처리
  퇴장·퇴장(강제퇴장)   → 구간 종료
순공부 = Σ(입장/재입장 → 다음 종료), 제외 = Σ(외출/이동 → 다음 재입장/종료)

반(수업) 정보는 로그·명부 모두에 없어 'YYYY년 M월 정규' 단일 반으로 둔다.
좌석·연락처·학년 등 프로필은 student-info.xlsx 명부에서 이름으로 매칭해 보강한다.
출력 스키마는 기존 web/data.js(window.STUDYCORE_DATA)와 동일하다.
"""
import io
import json
import math
import os
import sys
import datetime
from collections import defaultdict

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC_LOG = os.path.join(ROOT, "inout_log.xlsx")
SRC_ROSTER = os.path.join(ROOT, "student-info.xlsx")
OUT = os.path.join(ROOT, "web", "data.js")

IN_EVENTS = {"입장", "재입장"}
BREAK_EVENTS = {"외출", "이동"}          # 자리 비움 = 제외시간 ('이동'도 제외)
CLOSE_EVENTS = {"퇴장", "퇴장(강제퇴장)"}
FORCED = "퇴장(강제퇴장)"


# ---------------------------------------------------------------- 유틸
def to_dt(v):
    if isinstance(v, datetime.datetime):
        return v
    return datetime.datetime.strptime(str(v), "%Y-%m-%d %H:%M:%S")


def bizday(dt):
    """새벽 5시 이전 이벤트는 전날 영업일로 귀속(자정 넘긴 공부)."""
    d = dt - datetime.timedelta(days=1) if dt.hour < 5 else dt
    return d.date().isoformat()


def hms(dt):
    return dt.strftime("%H:%M:%S") if dt else None


def is_weekend(date_str):
    return datetime.date.fromisoformat(date_str).weekday() >= 5


def digits(s):
    return "".join(ch for ch in str(s) if ch.isdigit()) if s is not None else ""


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ("", "-") else s


# ---------------------------------------------------------------- 로드
def load_events():
    wb = openpyxl.load_workbook(SRC_LOG, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    evs = []
    phone_of = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1 or r[0] in (None, ""):
            continue
        name = str(r[0]).strip()
        evs.append((name, to_dt(r[4]), r[3]))
        phone_of.setdefault(name, digits(r[1]))
    return evs, phone_of


def load_roster():
    """이름 -> {seat, loginPhones[], profile{}}. 파일 없으면 빈 dict."""
    if not os.path.exists(SRC_ROSTER):
        return {}
    wb = openpyxl.load_workbook(SRC_ROSTER, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rec = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1 or r[0] in (None, ""):
            continue
        name = str(r[0]).strip()
        l4 = []
        for p in (r[1], r[2]):  # 학생·보호자 연락처 둘 다 로그인 허용
            d = digits(p)
            if len(d) >= 4 and d[-4:] not in l4:
                l4.append(d[-4:])
        seat = r[4]
        seat = int(seat) if isinstance(seat, int) else (
            int(seat) if isinstance(seat, str) and seat.isdigit() else None)
        rec[name] = {
            "seat": seat,
            "loginPhones": l4,
            "profile": {
                "status": clean(r[3]), "gender": clean(r[5]), "birth": clean(r[6]),
                "grade": clean(r[7]), "school": clean(r[8]),
                "enrolledAt": clean(r[11]), "leftAt": clean(r[12]),
            },
        }
    return rec


# ---------------------------------------------------------------- 하루 집계
def build_day(evs):
    """evs: [(dt, status)] (시간 오름차순). 기존 day 스키마 dict 반환."""
    sessions = []
    cur = None
    open_study = None  # 공부 구간 시작 시각
    open_break = None  # 제외 구간 시작 시각

    def new_session(t):
        return {"in": t, "out": None, "netSec": 0.0, "excludedSec": 0.0,
                "outings": 0, "forced": False, "provisional": False}

    for t, st in evs:
        if st in IN_EVENTS:
            if cur is not None and open_break is not None:
                cur["excludedSec"] += (t - open_break).total_seconds()
            open_break = None
            if cur is None:
                cur = new_session(t)
            open_study = t
        elif st in BREAK_EVENTS:
            if cur is None:                       # 외출/이동이 먼저 찍힌 이상치
                cur = new_session(t)
            if open_study is not None:
                cur["netSec"] += (t - open_study).total_seconds()
                open_study = None
            open_break = t
            cur["outings"] += 1
        elif st in CLOSE_EVENTS:
            if cur is None:
                cur = new_session(t)
            if open_study is not None:
                cur["netSec"] += (t - open_study).total_seconds()
                open_study = None
            if open_break is not None:
                cur["excludedSec"] += (t - open_break).total_seconds()
                open_break = None
            cur["out"] = t
            if st == FORCED:
                cur["forced"] = True
            sessions.append(cur)
            cur = None
        # 그 외 상태는 무시
    if cur is not None:                            # 닫는 퇴장 없음(미완결: 주로 추출 시점 잘림)
        cur["provisional"] = True                  # 마지막 열린 구간은 종료 시각 불명 → 미가산
        sessions.append(cur)

    # 세션 스키마 변환 + 일자 집계
    out_sessions = []
    d_net = d_tot = d_exc = d_out = 0
    g_net = g_tot = g_exc = g_out = 0
    no_checkout = False
    for s in sessions:
        net = int(round(s["netSec"]))
        exc = int(round(s["excludedSec"]))
        tot = net + exc
        prov = s["provisional"]
        out_sessions.append({
            "in": hms(s["in"]), "out": hms(s["out"]),
            "totalSec": tot, "netSec": net, "excludedSec": exc, "outings": s["outings"],
            "reason": "강제퇴장" if s["forced"] else ("미복귀" if prov else "학습시간 인정 완료"),
            "provisional": prov, "seat": None,
        })
        d_net += net; d_tot += tot; d_exc += exc; d_out += s["outings"]
        if prov:
            no_checkout = True
        else:
            g_net += net; g_tot += tot; g_exc += exc; g_out += s["outings"]

    first_in = next((x["in"] for x in out_sessions if x["in"]), None)
    last_out = next((x["out"] for x in reversed(out_sessions) if x["out"]), None)
    return {
        "netSec": d_net, "totalSec": d_tot, "excludedSec": d_exc, "outings": d_out,
        "goodNetSec": g_net, "goodTotalSec": g_tot, "goodExcludedSec": g_exc, "goodOutings": g_out,
        "firstIn": first_in, "lastOut": last_out,
        "attended": True, "noCheckout": no_checkout, "sessions": out_sessions,
    }


# ---------------------------------------------------------------- 빌드
def build():
    evs, phone_of = load_events()
    roster = load_roster()

    by_block = defaultdict(list)        # (name, bizday) -> [(dt, status)]
    open_days = defaultdict(set)        # month -> {bizday}
    for name, dt, st in evs:
        bd = bizday(dt)
        by_block[(name, bd)].append((dt, st))
        open_days[bd[:7]].add(bd)
    for k in by_block:
        by_block[k].sort()

    # name -> month -> {date: day}
    sd = defaultdict(lambda: defaultdict(dict))
    for (name, bd), block in by_block.items():
        sd[name][bd[:7]][bd] = build_day(block)

    students = []
    class_acc = defaultdict(lambda: defaultdict(list))
    for name in sorted(sd.keys()):
        months_out = {}
        for m, days in sd[name].items():
            total_net = att = rec = noco = 0
            wd_list, we_list = [], []
            max_day = {"date": None, "netSec": -1}
            for date in sorted(days.keys()):
                d = days[date]
                att += 1
                total_net += d["netSec"]
                if d["noCheckout"] and d["netSec"] == 0:
                    noco += 1
                if d["netSec"] > 0:
                    rec += 1
                    (we_list if is_weekend(date) else wd_list).append(d["netSec"])
                if d["netSec"] > max_day["netSec"]:
                    max_day = {"date": date, "netSec": d["netSec"]}
            daily = round(total_net / rec) if rec else 0
            wd_avg = round(sum(wd_list) / len(wd_list)) if wd_list else 0
            we_avg = round(sum(we_list) / len(we_list)) if we_list else 0
            months_out[m] = {
                "totalNetSec": total_net, "attendanceDays": att, "recognizedDays": rec,
                "noCheckoutDays": noco, "openDays": len(open_days[m]),
                "dailyAvgSec": daily, "weekdayAvgSec": wd_avg, "weekendAvgSec": we_avg,
                "maxDay": max_day if max_day["date"] else None,
                "className": f"{int(m[:4])}년 {int(m[5:7])}월 정규",
                "days": days,
            }
            if rec:
                class_acc[m]["totalNetSec"].append(total_net)
                class_acc[m]["attendanceDays"].append(att)
                class_acc[m]["dailyAvgSec"].append(daily)
                if wd_avg:
                    class_acc[m]["weekdayAvgSec"].append(wd_avg)
                if we_avg:
                    class_acc[m]["weekendAvgSec"].append(we_avg)

        # 명부 보강 (좌석·로그인전화·프로필)
        r = roster.get(name)
        seat = phone_last4 = profile = None
        login_phones = []
        if r:
            seat = r["seat"]
            login_phones = r["loginPhones"]
            phone_last4 = login_phones[0] if login_phones else None
            profile = r["profile"]
        if not phone_last4:                         # 명부에 없으면 로그의 실제 번호 뒷4
            p = phone_of.get(name, "")
            phone_last4 = p[-4:] if len(p) >= 4 else None

        st_obj = {"key": name, "name": name, "phoneLast4": phone_last4,
                  "seat": seat, "months": months_out}
        if login_phones:
            st_obj["loginPhones"] = login_phones
            st_obj["phoneFromRoster"] = True
        if profile:
            st_obj["profile"] = profile
        students.append(st_obj)

    # 반 평균(익명)
    def top_mean(vals, frac=0.2):
        if not vals:
            return 0
        s = sorted(vals, reverse=True)
        k = max(1, math.ceil(len(s) * frac))
        return round(sum(s[:k]) / k)

    class_averages = {}
    for m, acc in class_acc.items():
        n = len(acc["totalNetSec"])
        class_averages[m] = {
            "studentCount": n,
            "totalNetSec": round(sum(acc["totalNetSec"]) / n) if n else 0,
            "attendanceDays": round(sum(acc["attendanceDays"]) / n, 1) if n else 0,
            "dailyAvgSec": round(sum(acc["dailyAvgSec"]) / n) if n else 0,
            "weekdayAvgSec": round(sum(acc["weekdayAvgSec"]) / len(acc["weekdayAvgSec"]))
            if acc["weekdayAvgSec"] else 0,
            "weekendAvgSec": round(sum(acc["weekendAvgSec"]) / len(acc["weekendAvgSec"]))
            if acc["weekendAvgSec"] else 0,
            "top20Count": max(1, math.ceil(n * 0.2)) if n else 0,
            "top20TotalNetSec": top_mean(acc["totalNetSec"]),
            "top20DailyAvgSec": top_mean(acc["dailyAvgSec"]),
        }

    all_months = sorted(open_days.keys())
    return {
        "generatedFrom": "inout_log.xlsx",
        "months": all_months,
        "openDays": {m: sorted(open_days[m]) for m in all_months},
        "classAverages": class_averages,
        "students": students,
    }


def main():
    data = build()
    os.makedirs(os.path.join(ROOT, "web"), exist_ok=True)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// 자동 생성됨 (scripts/preprocess.py). 직접 수정 금지.\n")
        f.write("window.STUDYCORE_DATA = ")
        f.write(payload)
        f.write(";\n")

    print(f"[OK] {OUT}")
    print(f"  generatedFrom: {data['generatedFrom']}")
    print(f"  months: {data['months']}")
    print(f"  students: {len(data['students'])}")
    for m, ca in data["classAverages"].items():
        print(f"  반평균[{m}]: 학생 {ca['studentCount']}명, "
              f"총 {ca['totalNetSec'] // 3600}시간, 일평균 {ca['dailyAvgSec'] // 60}분")
    print("  로그인 예시 (이름 / 전화뒷4):")
    for s in data["students"][:8]:
        print(f"    {s['name']} / {s['phoneLast4']}")


if __name__ == "__main__":
    main()
